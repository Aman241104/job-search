import json
import uuid
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import DATA_DIR, OUTPUT_DIR, PROFILE_ENCRYPTION_KEY
from rich.console import Console
from rich.table import Table
from cryptography.fernet import Fernet, InvalidToken

console = Console()

_fernet = Fernet(PROFILE_ENCRYPTION_KEY.encode()) if PROFILE_ENCRYPTION_KEY else None


def encrypt_secret(plaintext: str) -> str:
    """Used for per-user secrets stored in the `profiles` JSONB blob (e.g.
    smtp_app_password) — these are real third-party credentials and must not
    sit in Postgres as plaintext. No-ops (stores as-is) if
    PROFILE_ENCRYPTION_KEY isn't set, so local dev without the key doesn't
    hard-fail — only degrades to the pre-encryption behavior."""
    if not plaintext or not _fernet:
        return plaintext
    return _fernet.encrypt(plaintext.encode()).decode()


def decrypt_secret(ciphertext: str) -> str:
    """Returns '' on any decrypt failure (wrong/rotated key, or a legacy
    plaintext value from before encryption existed) rather than raising —
    a broken decrypt must degrade to "not configured", not crash the caller
    (e.g. an email-send attempt)."""
    if not ciphertext or not _fernet:
        return ciphertext
    try:
        return _fernet.decrypt(ciphertext.encode()).decode()
    except (InvalidToken, ValueError):
        return ""

# ── Postgres-only (multi-tenant migration, 2026-07-23) ──────────────────────
# SQLite dropped entirely — every table is now scoped by user_id, and local
# dev needs the same real users table Postgres provides, so there's no more
# zero-setup single-user fallback. DATABASE_URL is required.
DATABASE_URL = os.getenv("DATABASE_URL", "")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL is required — SQLite fallback was removed in the multi-tenant migration")

import psycopg2
import psycopg2.pool
_pg_pool = psycopg2.pool.ThreadedConnectionPool(1, 10, DATABASE_URL)

ROW_DICT = True  # pass to `conn.row_factory = ROW_DICT` for dict-shaped rows


class _DictCursorProxy:
    """Wraps a plain psycopg2 cursor so fetchone()/fetchall() return dicts."""

    def __init__(self, cursor):
        self._cursor = cursor

    def _cols(self):
        return [c[0] for c in self._cursor.description]

    def fetchone(self):
        row = self._cursor.fetchone()
        return dict(zip(self._cols(), row)) if row is not None else None

    def fetchall(self):
        cols = self._cols()
        return [dict(zip(cols, row)) for row in self._cursor.fetchall()]

    @property
    def description(self):
        return self._cursor.description


class _PgConnWrapper:
    """Pooled-Postgres connection with a sqlite3-connection-shaped `.execute()`
    (accepts `?` placeholders, auto-translated to `%s`) so every existing
    caller in this file and in app.py keeps working unchanged. Returns the
    connection to the pool on __exit__ instead of closing it."""

    def __init__(self, pool):
        self._pool = pool
        self._conn = pool.getconn()
        self.row_factory = None

    def execute(self, sql, params=()):
        cursor = self._conn.cursor()
        cursor.execute(sql.replace("?", "%s"), params)
        return _DictCursorProxy(cursor) if self.row_factory else cursor

    def commit(self):
        self._conn.commit()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self._conn.commit()
        else:
            self._conn.rollback()
        self._pool.putconn(self._conn)


STATUS_COLORS = {
    "found": "D3D3D3",        # gray
    "applied": "ADD8E6",      # light blue
    "interviewing": "FFFF99", # yellow
    "offer": "90EE90",        # green
    "rejected": "FFB6C1",     # red/pink
    "ghosted": "E8E8E8",      # light gray
}


class TrackerAgent:
    def __init__(self):
        self.excel_path = Path(OUTPUT_DIR) / "job_tracker.xlsx"
        self._init_db()

    def _get_conn(self):
        return _PgConnWrapper(_pg_pool)

    def _init_db(self):
        with self._get_conn() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id TEXT PRIMARY KEY,
                    google_sub TEXT UNIQUE,
                    email TEXT,
                    name TEXT,
                    avatar_url TEXT,
                    created_at TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS jobs (
                    id TEXT PRIMARY KEY,
                    user_id TEXT REFERENCES users(id),
                    title TEXT,
                    company TEXT,
                    description TEXT,
                    url TEXT,
                    source TEXT,
                    location TEXT,
                    salary TEXT,
                    date_found TEXT,
                    date_posted TEXT,
                    tags TEXT,
                    score INTEGER DEFAULT 0,
                    score_reason TEXT,
                    starred INTEGER DEFAULT 0,
                    legitimacy_score INTEGER,
                    legitimacy_flags TEXT
                )
            """)
            conn.execute("ALTER TABLE jobs ADD COLUMN IF NOT EXISTS date_posted TEXT")
            # url uniqueness is per-user (two users can independently find the
            # same posting) — a bare UNIQUE(url) from the single-tenant schema
            # would let user B's scrape silently no-op against user A's row.
            conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_jobs_user_url ON jobs(user_id, url)")
            conn.execute("""
                CREATE TABLE IF NOT EXISTS applications (
                    id TEXT PRIMARY KEY,
                    user_id TEXT REFERENCES users(id),
                    job_id TEXT UNIQUE,
                    date_applied TEXT,
                    status TEXT DEFAULT 'found',
                    cv_path TEXT,
                    cover_letter_path TEXT,
                    cv_markdown TEXT,
                    cover_letter_text TEXT,
                    notes TEXT,
                    last_updated TEXT,
                    FOREIGN KEY (job_id) REFERENCES jobs(id)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS training_sessions (
                    id TEXT PRIMARY KEY,
                    user_id TEXT REFERENCES users(id),
                    topic_key TEXT,
                    topic_name TEXT,
                    messages TEXT,
                    avg_score REAL DEFAULT 0,
                    created_at TEXT,
                    last_updated TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS blacklisted_companies (
                    user_id TEXT REFERENCES users(id),
                    company TEXT,
                    blacklisted_at TEXT,
                    PRIMARY KEY (user_id, company)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS interview_rounds (
                    id TEXT PRIMARY KEY,
                    user_id TEXT REFERENCES users(id),
                    job_id TEXT,
                    round_num INTEGER,
                    round_type TEXT,
                    scheduled_at TEXT,
                    result TEXT,
                    notes TEXT,
                    created_at TEXT,
                    FOREIGN KEY (job_id) REFERENCES jobs(id)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS learning_items (
                    id TEXT PRIMARY KEY,
                    user_id TEXT REFERENCES users(id),
                    title TEXT,
                    item_type TEXT,
                    phase INTEGER,
                    order_index INTEGER,
                    status TEXT DEFAULT 'not_started',
                    notes TEXT,
                    updated_at TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS telegram_alerts (
                    message_id TEXT PRIMARY KEY,
                    user_id TEXT REFERENCES users(id),
                    job_id TEXT,
                    sent_at TEXT,
                    FOREIGN KEY (job_id) REFERENCES jobs(id)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS learning_topics (
                    id TEXT PRIMARY KEY,
                    item_id TEXT,
                    topic_name TEXT,
                    order_index INTEGER,
                    covered INTEGER DEFAULT 0,
                    FOREIGN KEY (item_id) REFERENCES learning_items(id)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS learning_books (
                    id TEXT PRIMARY KEY,
                    user_id TEXT REFERENCES users(id),
                    title TEXT,
                    filename TEXT,
                    page_count INTEGER DEFAULT 0,
                    current_page INTEGER DEFAULT 1,
                    uploaded_at TEXT,
                    cloudinary_url TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS learning_book_pages (
                    id TEXT PRIMARY KEY,
                    book_id TEXT,
                    page_num INTEGER,
                    text TEXT,
                    summary TEXT,
                    FOREIGN KEY (book_id) REFERENCES learning_books(id)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS story_bank (
                    id TEXT PRIMARY KEY,
                    user_id TEXT REFERENCES users(id),
                    situation TEXT,
                    task TEXT,
                    action TEXT,
                    result TEXT,
                    reflection TEXT,
                    tags TEXT,
                    source_job_id TEXT,
                    created_at TEXT,
                    updated_at TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS settings (
                    user_id TEXT REFERENCES users(id),
                    key TEXT,
                    value TEXT,
                    PRIMARY KEY (user_id, key)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS batches (
                    id TEXT PRIMARY KEY,
                    user_id TEXT REFERENCES users(id),
                    mode TEXT,
                    channel TEXT,
                    status TEXT DEFAULT 'staged',
                    created_at TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS batch_items (
                    id TEXT PRIMARY KEY,
                    batch_id TEXT,
                    job_id TEXT,
                    email TEXT,
                    cv_path TEXT,
                    cover_path TEXT,
                    cv_markdown TEXT,
                    cover_letter_text TEXT,
                    screenshot_url TEXT,
                    fields_filled TEXT,
                    fields_missing TEXT,
                    approved INTEGER DEFAULT 1,
                    status TEXT DEFAULT 'staged',
                    error TEXT,
                    FOREIGN KEY (batch_id) REFERENCES batches(id),
                    FOREIGN KEY (job_id) REFERENCES jobs(id)
                )
            """)
            conn.execute("CREATE INDEX IF NOT EXISTS idx_batch_items_batch ON batch_items(batch_id)")
            conn.execute("""
                CREATE TABLE IF NOT EXISTS learning_playlists (
                    id TEXT PRIMARY KEY,
                    user_id TEXT REFERENCES users(id),
                    url TEXT,
                    title TEXT,
                    status TEXT DEFAULT 'ingesting',
                    created_at TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS learning_videos (
                    id TEXT PRIMARY KEY,
                    playlist_id TEXT,
                    video_id TEXT,
                    title TEXT,
                    url TEXT,
                    transcript TEXT,
                    notes_md TEXT,
                    source TEXT,
                    status TEXT DEFAULT 'pending',
                    error TEXT,
                    created_at TEXT,
                    FOREIGN KEY (playlist_id) REFERENCES learning_playlists(id)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS learning_video_chunks (
                    id TEXT PRIMARY KEY,
                    video_id TEXT,
                    playlist_id TEXT,
                    chunk_index INTEGER,
                    text TEXT,
                    created_at TEXT,
                    FOREIGN KEY (video_id) REFERENCES learning_videos(id)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS resumes (
                    user_id TEXT PRIMARY KEY REFERENCES users(id),
                    data JSONB,
                    updated_at TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS profiles (
                    user_id TEXT PRIMARY KEY REFERENCES users(id),
                    data JSONB,
                    updated_at TEXT
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS rate_limit_hits (
                    user_id TEXT NOT NULL,
                    action TEXT NOT NULL,
                    ts TEXT NOT NULL
                )
            """)
            conn.execute("CREATE INDEX IF NOT EXISTS idx_rate_limit_lookup ON rate_limit_hits(user_id, action, ts)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_learning_topics_item ON learning_topics(item_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_book_pages_book ON learning_book_pages(book_id)")
            # Indexes for performance + per-user scoping
            conn.execute("CREATE INDEX IF NOT EXISTS idx_jobs_user ON jobs(user_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_jobs_score ON jobs(score DESC)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_jobs_source ON jobs(source)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_jobs_date ON jobs(date_found DESC)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_jobs_starred ON jobs(starred)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_apps_user ON applications(user_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_apps_status ON applications(status)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_apps_job_id ON applications(job_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_training_user ON training_sessions(user_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_interview_user ON interview_rounds(user_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_learning_items_user ON learning_items(user_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_learning_books_user ON learning_books(user_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_story_bank_user ON story_bank(user_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_batches_user ON batches(user_id)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_playlists_user ON learning_playlists(user_id)")
            conn.commit()

    # ── Users (Google OAuth login — see auth.py) ────────────────────────────────

    def get_or_create_user(self, google_sub: str, email: str, name: str, avatar_url: str) -> str:
        """Returns the user's id, creating a row on first login."""
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute("SELECT id FROM users WHERE google_sub = ?", (google_sub,)).fetchone()
            if row:
                return row["id"]

            user_id = str(uuid.uuid4())
            conn.execute(
                "INSERT INTO users (id, google_sub, email, name, avatar_url, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (user_id, google_sub, email, name, avatar_url, datetime.now().isoformat()),
            )
            conn.commit()
        return user_id

    def get_user_by_email(self, email: str) -> dict | None:
        """For non-cookie callers (e.g. the Telegram webhook — Telegram itself
        calls it, no browser session) that need to resolve to a specific
        account by a stable identifier other than the session cookie."""
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute("SELECT id, email, name, avatar_url, created_at FROM users WHERE email = ?", (email,)).fetchone()
        return row

    def get_user_by_telegram_chat_id(self, chat_id: str) -> dict | None:
        """For the Telegram webhook — Telegram calls this directly with no
        browser session, so an inbound message is routed to a user by
        matching the chat_id captured during that user's own connect-link
        flow (see /api/telegram/connect-link + webhook in app.py)."""
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute(
                "SELECT user_id FROM profiles WHERE data->>'telegram_chat_id' = ?", (str(chat_id),)
            ).fetchone()
        return self.get_user(row["user_id"]) if row else None

    def get_user_by_telegram_token(self, token: str) -> dict | None:
        """Resolves the one-time /start <token> deep-link payload back to the
        account that generated it, so the webhook can link that chat_id to
        the right user on first contact."""
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute(
                "SELECT user_id FROM profiles WHERE data->>'telegram_connect_token' = ?", (token,)
            ).fetchone()
        return self.get_user(row["user_id"]) if row else None

    def get_user(self, user_id: str) -> dict | None:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute("SELECT id, email, name, avatar_url, created_at FROM users WHERE id = ?", (user_id,)).fetchone()
        return row

    # ── Resume + profile (one JSONB blob per user) ──────────────────────────────

    def get_resume(self, user_id: str) -> dict | None:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute("SELECT data FROM resumes WHERE user_id = ?", (user_id,)).fetchone()
        return row["data"] if row else None

    def save_resume(self, user_id: str, data: dict) -> None:
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute("""
                INSERT INTO resumes (user_id, data, updated_at) VALUES (?, ?, ?)
                ON CONFLICT (user_id) DO UPDATE SET data = EXCLUDED.data, updated_at = EXCLUDED.updated_at
            """, (user_id, json.dumps(data), now))
            conn.commit()

    def get_profile(self, user_id: str) -> dict | None:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute("SELECT data FROM profiles WHERE user_id = ?", (user_id,)).fetchone()
        return row["data"] if row else None

    def save_profile(self, user_id: str, data: dict) -> None:
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute("""
                INSERT INTO profiles (user_id, data, updated_at) VALUES (?, ?, ?)
                ON CONFLICT (user_id) DO UPDATE SET data = EXCLUDED.data, updated_at = EXCLUDED.updated_at
            """, (user_id, json.dumps(data), now))
            conn.commit()

    # ── Account deletion + full data export ─────────────────────────────────────

    def delete_account(self, user_id: str) -> None:
        """Cascades every user-scoped table in one transaction — none of
        these FKs have ON DELETE CASCADE, so children must go before
        parents. Cloudinary-hosted files (book/resume PDFs) are NOT purged
        here, only DB rows — a real but accepted gap for this first pass."""
        with self._get_conn() as conn:
            conn.execute(
                "DELETE FROM learning_video_chunks WHERE video_id IN "
                "(SELECT id FROM learning_videos WHERE playlist_id IN "
                "(SELECT id FROM learning_playlists WHERE user_id = ?))", (user_id,))
            conn.execute(
                "DELETE FROM learning_videos WHERE playlist_id IN "
                "(SELECT id FROM learning_playlists WHERE user_id = ?)", (user_id,))
            conn.execute("DELETE FROM learning_playlists WHERE user_id = ?", (user_id,))
            conn.execute(
                "DELETE FROM learning_book_pages WHERE book_id IN "
                "(SELECT id FROM learning_books WHERE user_id = ?)", (user_id,))
            conn.execute("DELETE FROM learning_books WHERE user_id = ?", (user_id,))
            conn.execute(
                "DELETE FROM learning_topics WHERE item_id IN "
                "(SELECT id FROM learning_items WHERE user_id = ?)", (user_id,))
            conn.execute("DELETE FROM learning_items WHERE user_id = ?", (user_id,))
            conn.execute(
                "DELETE FROM batch_items WHERE batch_id IN "
                "(SELECT id FROM batches WHERE user_id = ?)", (user_id,))
            conn.execute("DELETE FROM batches WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM telegram_alerts WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM interview_rounds WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM applications WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM jobs WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM training_sessions WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM blacklisted_companies WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM story_bank WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM settings WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM resumes WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM profiles WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM rate_limit_hits WHERE user_id = ?", (user_id,))
            conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
            conn.commit()

    def export_user_data(self, user_id: str) -> dict:
        """Full 'download my data' dump — every user-scoped table as JSON.
        The stored smtp_app_password is encrypted ciphertext, not a value
        meant to leave the server even in an export a user requested
        themselves — replaced with the same boolean flag GET /api/user/profile
        uses. Cloudinary-hosted files (resume/book PDFs) are referenced only
        by their stored path/URL, not re-downloaded into this export."""
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT

            def rows(sql):
                return conn.execute(sql, (user_id,)).fetchall()

            user = conn.execute(
                "SELECT id, email, name, avatar_url, created_at FROM users WHERE id = ?", (user_id,)
            ).fetchone()
            profile_row = conn.execute("SELECT data FROM profiles WHERE user_id = ?", (user_id,)).fetchone()
            profile = dict(profile_row["data"]) if profile_row else {}
            if profile.get("smtp_app_password"):
                profile["smtp_app_password"] = None
                profile["smtp_app_password_set"] = True
            resume_row = conn.execute("SELECT data FROM resumes WHERE user_id = ?", (user_id,)).fetchone()

            return {
                "user": user,
                "profile": profile,
                "resume": resume_row["data"] if resume_row else None,
                "jobs": rows("SELECT * FROM jobs WHERE user_id = ?"),
                "applications": rows("SELECT * FROM applications WHERE user_id = ?"),
                "training_sessions": rows("SELECT * FROM training_sessions WHERE user_id = ?"),
                "interview_rounds": rows("SELECT * FROM interview_rounds WHERE user_id = ?"),
                "story_bank": rows("SELECT * FROM story_bank WHERE user_id = ?"),
                "blacklisted_companies": rows("SELECT * FROM blacklisted_companies WHERE user_id = ?"),
                "batches": rows("SELECT * FROM batches WHERE user_id = ?"),
                "learning_items": rows("SELECT * FROM learning_items WHERE user_id = ?"),
                "learning_books": rows("SELECT * FROM learning_books WHERE user_id = ?"),
                "learning_playlists": rows("SELECT * FROM learning_playlists WHERE user_id = ?"),
            }

    def get_auto_find_user_ids(self) -> list:
        """Users who opted into the daily auto-find cron (see
        app.py's /internal/cron/auto-find) — opt-in, not opt-on-by-default,
        so nobody gets automatic scraping/Telegram alerts they didn't ask for."""
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute(
                "SELECT user_id FROM profiles WHERE data->>'auto_find_enabled' = 'true'"
            ).fetchall()
        return [r["user_id"] for r in rows]

    # ── Rate limiting (per-user, per-action sliding window) ─────────────────────

    def check_rate_limit(self, user_id: str, action: str, max_calls: int, window_seconds: int) -> bool:
        """Returns True and records the hit if the caller is still under
        `max_calls` within the trailing `window_seconds`; returns False
        (does not record) if they're over. Postgres-backed rather than an
        in-memory counter so the limit holds even if Cloud Run scales to
        multiple instances."""
        cutoff = (datetime.now() - timedelta(seconds=window_seconds)).isoformat()
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            # opportunistic cleanup — keeps this table from growing unbounded
            # without needing a separate cron job
            conn.execute(
                "DELETE FROM rate_limit_hits WHERE user_id = ? AND action = ? AND ts <= ?",
                (user_id, action, cutoff),
            )
            row = conn.execute(
                "SELECT COUNT(*) as c FROM rate_limit_hits WHERE user_id = ? AND action = ? AND ts > ?",
                (user_id, action, cutoff),
            ).fetchone()
            if row["c"] >= max_calls:
                return False
            conn.execute(
                "INSERT INTO rate_limit_hits (user_id, action, ts) VALUES (?, ?, ?)",
                (user_id, action, datetime.now().isoformat()),
            )
            conn.commit()
        return True

    # ── Jobs + applications ─────────────────────────────────────────────────────

    def add_job(self, user_id: str, job: dict) -> bool:
        try:
            with self._get_conn() as conn:
                # RETURNING id gives us the row actually written — nothing comes
                # back when ON CONFLICT DO NOTHING skipped the insert (same URL
                # already scraped for this user under a different generated UUID).
                cursor = conn.execute("""
                    INSERT INTO jobs
                    (id, user_id, title, company, description, url, source, location, salary, date_found, date_posted, tags, score, score_reason)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT (user_id, url) DO NOTHING
                    RETURNING id
                """, (
                    job["id"], user_id, job["title"], job["company"], job.get("description", ""),
                    job["url"], job.get("source", ""), job.get("location", ""),
                    job.get("salary", ""), job.get("date_found", datetime.now().isoformat()),
                    job.get("date_posted", ""),
                    json.dumps(job.get("tags", [])), job.get("score", 0), job.get("score_reason", "")
                ))
                row = cursor.fetchone()
                job_id = row[0] if row else None

                if job_id is None:
                    existing = conn.execute(
                        "SELECT id FROM jobs WHERE user_id = ? AND url = ?", (user_id, job["url"])
                    ).fetchone()
                    if not existing:
                        return False
                    job_id = existing[0]

                conn.execute("""
                    INSERT INTO applications (id, user_id, job_id, status, last_updated)
                    VALUES (?, ?, ?, 'found', ?)
                    ON CONFLICT (job_id) DO NOTHING
                """, (str(uuid.uuid4()), user_id, job_id, datetime.now().isoformat()))
                conn.commit()
            return True
        except Exception as e:
            console.print(f"[red]DB error adding job: {e}[/red]")
            return False

    def update_status(self, user_id: str, job_id: str, status: str, notes: str = "", cv_path: str = "", cover_path: str = ""):
        valid = {"found", "applied", "interviewing", "offer", "rejected", "ghosted", "skipped"}
        if status not in valid:
            console.print(f"[red]Invalid status. Choose from: {valid}[/red]")
            return
        with self._get_conn() as conn:
            now = datetime.now().isoformat()
            date_applied = now if status == "applied" else None
            conn.execute("""
                UPDATE applications SET status=?, notes=?, last_updated=?,
                cv_path=COALESCE(NULLIF(?, ''), cv_path),
                cover_letter_path=COALESCE(NULLIF(?, ''), cover_letter_path),
                date_applied=COALESCE(NULLIF(?, ''), date_applied)
                WHERE job_id=? AND user_id=?
            """, (status, notes, now, cv_path, cover_path, date_applied or '', job_id, user_id))
            conn.commit()
        console.print(f"[green]Status updated to '{status}'[/green]")

    def toggle_star(self, user_id: str, job_id: str) -> bool:
        with self._get_conn() as conn:
            conn.execute(
                "UPDATE jobs SET starred = 1 - COALESCE(starred, 0) WHERE id = ? AND user_id = ?",
                (job_id, user_id)
            )
            conn.commit()
            row = conn.execute("SELECT starred FROM jobs WHERE id = ? AND user_id = ?", (job_id, user_id)).fetchone()
            return bool(row[0]) if row else False

    def get_legitimacy(self, user_id: str, job_id: str) -> dict | None:
        """Returns None if never computed (caller should compute + save)."""
        with self._get_conn() as conn:
            row = conn.execute(
                "SELECT legitimacy_score, legitimacy_flags FROM jobs WHERE id = ? AND user_id = ?", (job_id, user_id)
            ).fetchone()
        if not row or row[0] is None:
            return None
        try:
            flags = json.loads(row[1]) if row[1] else []
        except Exception:
            flags = []
        return {"score": row[0], "flags": flags}

    def save_legitimacy(self, user_id: str, job_id: str, score: int, flags: list) -> None:
        with self._get_conn() as conn:
            conn.execute(
                "UPDATE jobs SET legitimacy_score = ?, legitimacy_flags = ? WHERE id = ? AND user_id = ?",
                (score, json.dumps(flags), job_id, user_id),
            )
            conn.commit()

    # ── Interview Story Bank (STAR + Reflection) ────────────────────────────────

    def add_story(self, user_id: str, situation: str, task: str, action: str, result: str,
                  reflection: str, tags: list, source_job_id: str = "") -> str:
        story_id = str(uuid.uuid4())
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute("""
                INSERT INTO story_bank
                (id, user_id, situation, task, action, result, reflection, tags, source_job_id, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (story_id, user_id, situation, task, action, result, reflection,
                  json.dumps(tags), source_job_id or None, now, now))
            conn.commit()
        return story_id

    def get_stories(self, user_id: str) -> list:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute(
                "SELECT * FROM story_bank WHERE user_id = ? ORDER BY updated_at DESC", (user_id,)
            ).fetchall()
        stories = []
        for d in rows:
            try:
                d['tags'] = json.loads(d['tags']) if d.get('tags') else []
            except Exception:
                d['tags'] = []
            stories.append(d)
        return stories

    def update_story(self, user_id: str, story_id: str, situation: str, task: str, action: str,
                      result: str, reflection: str, tags: list) -> None:
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute("""
                UPDATE story_bank
                SET situation=?, task=?, action=?, result=?, reflection=?, tags=?, updated_at=?
                WHERE id=? AND user_id=?
            """, (situation, task, action, result, reflection, json.dumps(tags), now, story_id, user_id))
            conn.commit()

    def delete_story(self, user_id: str, story_id: str) -> None:
        with self._get_conn() as conn:
            conn.execute("DELETE FROM story_bank WHERE id = ? AND user_id = ?", (story_id, user_id))
            conn.commit()

    # ── Settings (simple key-value) ─────────────────────────────────────────────

    def get_setting(self, user_id: str, key: str, default: str = "") -> str:
        with self._get_conn() as conn:
            row = conn.execute("SELECT value FROM settings WHERE user_id = ? AND key = ?", (user_id, key)).fetchone()
        if not row:
            return default
        return row[0] if not isinstance(row, dict) else row["value"]

    def set_setting(self, user_id: str, key: str, value: str) -> None:
        with self._get_conn() as conn:
            conn.execute("""
                INSERT INTO settings (user_id, key, value) VALUES (?, ?, ?)
                ON CONFLICT (user_id, key) DO UPDATE SET value = EXCLUDED.value
            """, (user_id, key, value))
            conn.commit()

    # ── Batch apply (email/browser, automatic or review-then-send) ─────────────

    def create_batch(self, user_id: str, mode: str, channel: str) -> str:
        batch_id = str(uuid.uuid4())
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute(
                "INSERT INTO batches (id, user_id, mode, channel, status, created_at) VALUES (?, ?, ?, ?, 'staged', ?)",
                (batch_id, user_id, mode, channel, now),
            )
            conn.commit()
        return batch_id

    def add_batch_item(self, batch_id: str, job_id: str, **fields) -> str:
        item_id = str(uuid.uuid4())
        cols = ["id", "batch_id", "job_id"] + list(fields.keys())
        placeholders = ", ".join(["?"] * len(cols))
        values = [item_id, batch_id, job_id] + list(fields.values())
        with self._get_conn() as conn:
            conn.execute(f"INSERT INTO batch_items ({', '.join(cols)}) VALUES ({placeholders})", values)
            conn.commit()
        return item_id

    def get_batch(self, user_id: str, batch_id: str) -> dict | None:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            batch_row = conn.execute(
                "SELECT * FROM batches WHERE id = ? AND user_id = ?", (batch_id, user_id)
            ).fetchone()
            if not batch_row:
                return None
            item_rows = conn.execute("""
                SELECT bi.*, j.title, j.company, j.score
                FROM batch_items bi JOIN jobs j ON bi.job_id = j.id
                WHERE bi.batch_id = ?
            """, (batch_id,)).fetchall()
        batch = batch_row
        batch["items"] = item_rows
        return batch

    def set_batch_item_approval(self, item_id: str, approved: bool) -> None:
        with self._get_conn() as conn:
            conn.execute("UPDATE batch_items SET approved = ? WHERE id = ?", (1 if approved else 0, item_id))
            conn.commit()

    def update_batch_item_status(self, item_id: str, status: str, error: str = "") -> None:
        with self._get_conn() as conn:
            conn.execute("UPDATE batch_items SET status = ?, error = ? WHERE id = ?", (status, error, item_id))
            conn.commit()

    def update_batch_status(self, batch_id: str, status: str) -> None:
        with self._get_conn() as conn:
            conn.execute("UPDATE batches SET status = ? WHERE id = ?", (status, batch_id))
            conn.commit()

    # ── Training Sessions ──────────────────────────────────────────────────────

    def save_training_session(self, user_id: str, session_id: str, topic_key: str, topic_name: str,
                               messages: list, avg_score: float):
        now = datetime.now().isoformat()
        messages_json = json.dumps(messages)
        with self._get_conn() as conn:
            existing = conn.execute(
                "SELECT id FROM training_sessions WHERE id = ? AND user_id = ?", (session_id, user_id)
            ).fetchone()
            if existing:
                conn.execute("""
                    UPDATE training_sessions
                    SET topic_key=?, topic_name=?, messages=?, avg_score=?, last_updated=?
                    WHERE id=? AND user_id=?
                """, (topic_key, topic_name, messages_json, avg_score, now, session_id, user_id))
            else:
                conn.execute("""
                    INSERT INTO training_sessions
                    (id, user_id, topic_key, topic_name, messages, avg_score, created_at, last_updated)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (session_id, user_id, topic_key, topic_name, messages_json, avg_score, now, now))
            conn.commit()

    def get_training_session(self, user_id: str, session_id: str) -> dict | None:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute(
                "SELECT * FROM training_sessions WHERE id = ? AND user_id = ?", (session_id, user_id)
            ).fetchone()
        if not row:
            return None
        try:
            row['messages'] = json.loads(row['messages'] or '[]')
        except Exception:
            row['messages'] = []
        return row

    def get_training_progress(self, user_id: str) -> dict:
        with self._get_conn() as conn:
            count_row = conn.execute(
                "SELECT COUNT(DISTINCT id) FROM training_sessions WHERE user_id = ?", (user_id,)
            ).fetchone()
            sessions_completed = count_row[0] if count_row else 0

            if sessions_completed == 0:
                return {
                    'sessions_completed': 0,
                    'avg_score': 0,
                    'topics_covered': [],
                    'total_messages': 0,
                }

            avg_row = conn.execute(
                "SELECT AVG(avg_score) FROM training_sessions WHERE user_id = ?", (user_id,)
            ).fetchone()
            avg_score = round(avg_row[0], 1) if avg_row and avg_row[0] is not None else 0

            topic_rows = conn.execute(
                "SELECT DISTINCT topic_key FROM training_sessions WHERE user_id = ? AND topic_key IS NOT NULL", (user_id,)
            ).fetchall()
            topics_covered = [r[0] for r in topic_rows]

            msg_rows = conn.execute("SELECT messages FROM training_sessions WHERE user_id = ?", (user_id,)).fetchall()
            total_messages = 0
            for (msg_json,) in msg_rows:
                try:
                    total_messages += len(json.loads(msg_json or '[]'))
                except Exception:
                    pass

        return {
            'sessions_completed': sessions_completed,
            'avg_score': avg_score,
            'topics_covered': topics_covered,
            'total_messages': total_messages,
        }

    # ── Company Blacklist ──────────────────────────────────────────────────────

    def toggle_blacklist(self, user_id: str, company: str) -> bool:
        """Adds company if not present, removes if present. Returns True if now blacklisted."""
        with self._get_conn() as conn:
            existing = conn.execute(
                "SELECT company FROM blacklisted_companies WHERE user_id = ? AND company = ?", (user_id, company)
            ).fetchone()
            if existing:
                conn.execute(
                    "DELETE FROM blacklisted_companies WHERE user_id = ? AND company = ?", (user_id, company)
                )
                conn.commit()
                return False
            else:
                conn.execute(
                    "INSERT INTO blacklisted_companies (user_id, company, blacklisted_at) VALUES (?, ?, ?)",
                    (user_id, company, datetime.now().isoformat())
                )
                conn.commit()
                return True

    def get_blacklisted(self, user_id: str) -> list:
        with self._get_conn() as conn:
            rows = conn.execute(
                "SELECT company FROM blacklisted_companies WHERE user_id = ? ORDER BY company", (user_id,)
            ).fetchall()
        return [r[0] for r in rows]

    def is_blacklisted(self, user_id: str, company: str) -> bool:
        with self._get_conn() as conn:
            row = conn.execute(
                "SELECT company FROM blacklisted_companies WHERE user_id = ? AND company = ?", (user_id, company)
            ).fetchone()
        return row is not None

    # ── Interview Rounds ───────────────────────────────────────────────────────

    def add_interview_round(self, user_id: str, job_id: str, round_type: str,
                             scheduled_at: str = None, notes: str = None) -> dict:
        with self._get_conn() as conn:
            existing = conn.execute(
                "SELECT COUNT(*) FROM interview_rounds WHERE job_id = ? AND user_id = ?", (job_id, user_id)
            ).fetchone()
            round_num = (existing[0] or 0) + 1
            round_id = str(uuid.uuid4())
            now = datetime.now().isoformat()
            conn.execute("""
                INSERT INTO interview_rounds
                (id, user_id, job_id, round_num, round_type, scheduled_at, result, notes, created_at)
                VALUES (?, ?, ?, ?, ?, ?, 'pending', ?, ?)
            """, (round_id, user_id, job_id, round_num, round_type, scheduled_at, notes, now))
            conn.commit()
        return {
            'id': round_id,
            'job_id': job_id,
            'round_num': round_num,
            'round_type': round_type,
            'scheduled_at': scheduled_at,
            'result': 'pending',
            'notes': notes,
            'created_at': now,
        }

    def update_interview_round(self, user_id: str, round_id: str, result: str = None,
                                notes: str = None, scheduled_at: str = None) -> dict:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute(
                "SELECT * FROM interview_rounds WHERE id = ? AND user_id = ?", (round_id, user_id)
            ).fetchone()
            if not row:
                return {}
            new_result = result if result is not None else row.get('result')
            new_notes = notes if notes is not None else row.get('notes')
            new_scheduled = scheduled_at if scheduled_at is not None else row.get('scheduled_at')
            conn.execute("""
                UPDATE interview_rounds
                SET result=?, notes=?, scheduled_at=?
                WHERE id=? AND user_id=?
            """, (new_result, new_notes, new_scheduled, round_id, user_id))
            conn.commit()
            updated = conn.execute(
                "SELECT * FROM interview_rounds WHERE id = ? AND user_id = ?", (round_id, user_id)
            ).fetchone()
        return updated or {}

    def get_interview_rounds(self, user_id: str, job_id: str) -> list:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute(
                "SELECT * FROM interview_rounds WHERE job_id = ? AND user_id = ? ORDER BY round_num ASC",
                (job_id, user_id)
            ).fetchall()
        return rows

    def delete_interview_round(self, user_id: str, round_id: str):
        with self._get_conn() as conn:
            conn.execute("DELETE FROM interview_rounds WHERE id = ? AND user_id = ?", (round_id, user_id))
            conn.commit()

    def get_all_applications(self, user_id: str) -> list:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute("""
                SELECT j.id, j.title, j.company, j.url, j.source, j.location, j.salary,
                       j.date_found, j.date_posted, j.score, j.score_reason, j.starred,
                       a.status, a.date_applied, a.notes, a.last_updated, a.cv_path, a.cover_letter_path
                FROM jobs j
                LEFT JOIN applications a ON j.id = a.job_id
                WHERE j.user_id = ?
                ORDER BY j.score DESC, a.last_updated DESC
            """, (user_id,)).fetchall()
        return rows

    def get_unapplied_top_jobs(self, user_id: str, min_score: int = 60, limit: int = 10) -> list:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute("""
                SELECT j.*, a.status FROM jobs j
                LEFT JOIN applications a ON j.id = a.job_id
                WHERE j.user_id = ? AND (a.status = 'found' OR a.status IS NULL) AND j.score >= ?
                ORDER BY j.score DESC LIMIT ?
            """, (user_id, min_score, limit)).fetchall()
        return rows

    def get_stats(self, user_id: str) -> dict:
        with self._get_conn() as conn:
            rows = conn.execute(
                "SELECT status, COUNT(*) as count FROM applications WHERE user_id = ? GROUP BY status", (user_id,)
            ).fetchall()
            total = conn.execute("SELECT COUNT(*) FROM jobs WHERE user_id = ?", (user_id,)).fetchone()[0]
        stats = {"total": total}
        for row in rows:
            stats[row[0]] = row[1]
        return stats

    def export_to_excel(self, user_id: str):
        apps = self.get_all_applications(user_id)
        stats = self.get_stats(user_id)

        wb = openpyxl.Workbook()

        # Sheet 1: Applications
        ws1 = wb.active
        ws1.title = "Applications"

        headers = ["#", "Title", "Company", "Status", "Score", "Source", "Location",
                   "Salary", "Date Found", "Date Applied", "Notes", "URL", "Score Reason"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2E4057")

        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx, app in enumerate(apps, 2):
            status = app.get("status", "found")
            fill_color = STATUS_COLORS.get(status, "FFFFFF")
            row_fill = PatternFill("solid", fgColor=fill_color)

            values = [
                row_idx - 1,
                app.get("title", ""),
                app.get("company", ""),
                status.upper(),
                app.get("score", 0),
                app.get("source", ""),
                app.get("location", ""),
                app.get("salary", ""),
                app.get("date_found", "")[:10] if app.get("date_found") else "",
                app.get("date_applied", "")[:10] if app.get("date_applied") else "",
                app.get("notes", ""),
                app.get("url", ""),
                app.get("score_reason", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws1.cell(row=row_idx, column=col, value=val)
                cell.fill = row_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Auto-width
        col_widths = [5, 30, 20, 12, 8, 12, 15, 15, 12, 12, 25, 40, 30]
        for i, w in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        ws1.freeze_panes = "A2"

        # Sheet 2: Stats
        ws2 = wb.create_sheet("Stats")
        ws2.append(["Metric", "Count"])
        ws2["A1"].font = Font(bold=True)
        ws2["B1"].font = Font(bold=True)
        ws2.append(["Total Jobs Found", stats.get("total", 0)])
        ws2.append(["Applied", stats.get("applied", 0)])
        ws2.append(["Interviewing", stats.get("interviewing", 0)])
        ws2.append(["Offers", stats.get("offer", 0)])
        ws2.append(["Rejected", stats.get("rejected", 0)])
        ws2.append(["Ghosted", stats.get("ghosted", 0)])
        ws2.append(["Not Yet Applied (found)", stats.get("found", 0)])
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 10

        # Sheet 3: High Priority
        ws3 = wb.create_sheet("High Priority")
        priority_headers = ["Title", "Company", "Score", "Location", "URL", "Reason"]
        for col, h in enumerate(priority_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        priority_jobs = [a for a in apps if a.get("score", 0) >= 75 and a.get("status") == "found"]
        for row_idx, j in enumerate(priority_jobs, 2):
            ws3.cell(row=row_idx, column=1, value=j.get("title", ""))
            ws3.cell(row=row_idx, column=2, value=j.get("company", ""))
            ws3.cell(row=row_idx, column=3, value=j.get("score", 0))
            ws3.cell(row=row_idx, column=4, value=j.get("location", ""))
            ws3.cell(row=row_idx, column=5, value=j.get("url", ""))
            ws3.cell(row=row_idx, column=6, value=j.get("score_reason", ""))

        for col_dims in [("A", 30), ("B", 20), ("C", 8), ("D", 15), ("E", 40), ("F", 35)]:
            ws3.column_dimensions[col_dims[0]].width = col_dims[1]

        wb.save(self.excel_path)
        console.print(f"[green]Excel exported to: {self.excel_path}[/green]")
        return str(self.excel_path)

    def print_dashboard(self, user_id: str):
        stats = self.get_stats(user_id)
        top_jobs = self.get_unapplied_top_jobs(user_id, limit=10)

        # Stats table
        t = Table(title="[bold]Application Stats[/bold]", show_header=True)
        t.add_column("Status", style="bold")
        t.add_column("Count", justify="right")
        t.add_row("Total Found", str(stats.get("total", 0)))
        t.add_row("[blue]Applied[/blue]", str(stats.get("applied", 0)))
        t.add_row("[yellow]Interviewing[/yellow]", str(stats.get("interviewing", 0)))
        t.add_row("[green]Offers[/green]", str(stats.get("offer", 0)))
        t.add_row("[red]Rejected[/red]", str(stats.get("rejected", 0)))
        console.print(t)

        # Top unapplied jobs
        if top_jobs:
            t2 = Table(title="[bold]Top Jobs to Apply[/bold]", show_header=True)
            t2.add_column("Score", justify="right", style="green")
            t2.add_column("Title")
            t2.add_column("Company")
            t2.add_column("Location")
            t2.add_column("ID", style="dim")
            for j in top_jobs[:8]:
                t2.add_row(
                    str(j.get("score", 0)),
                    j.get("title", ""),
                    j.get("company", ""),
                    j.get("location", ""),
                    j.get("id", "")[:8] + "..."
                )
            console.print(t2)

    # ── Learning track (ROADMAP.md Learning Core Track — no new scope beyond
    # tracking status + an AI tutor chat per item; no PDF/book upload) ─────────

    def seed_learning_items(self, user_id: str, items: list):
        """Insert each seed item if it doesn't already exist for this user —
        never overwrites a status the user has already set, safe to call on
        every startup."""
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            for item in items:
                conn.execute("""
                    INSERT INTO learning_items (id, user_id, title, item_type, phase, order_index, status, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, 'not_started', ?)
                    ON CONFLICT (id) DO NOTHING
                """, (f"{user_id}:{item['id']}", user_id, item["title"], item["type"], item["phase"], item["order"], now))
            conn.commit()

    def get_learning_items(self, user_id: str) -> list:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute(
                "SELECT * FROM learning_items WHERE user_id = ? ORDER BY phase ASC, order_index ASC", (user_id,)
            ).fetchall()
        return rows

    def update_learning_status(self, user_id: str, item_id: str, status: str, notes: str = "") -> bool:
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute(
                "UPDATE learning_items SET status=?, notes=?, updated_at=? WHERE id=? AND user_id=?",
                (status, notes, now, item_id, user_id),
            )
            conn.commit()
        return True

    def add_custom_learning_item(self, user_id: str, item_id: str, title: str) -> None:
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute("""
                INSERT INTO learning_items (id, user_id, title, item_type, phase, order_index, status, updated_at)
                VALUES (?, ?, ?, 'skill', 0, 0, 'not_started', ?)
                ON CONFLICT (id) DO NOTHING
            """, (item_id, user_id, title, now))
            conn.commit()

    # ── Topic checklists (per learning item — curated book/course OR custom
    # skill — an AI-generated zero-to-hero breakdown, checked off manually for
    # a real 0-100 coverage score rather than fragile auto-detection). Scoped
    # by item_id, which is itself user-scoped above, so no separate user_id
    # column needed here. ────────────────────────────────────────────────────

    def save_learning_topics(self, item_id: str, topic_names: list) -> None:
        """Only inserts if this item has no topics yet — never regenerates
        over a checklist the user has already started checking off."""
        with self._get_conn() as conn:
            existing = conn.execute(
                "SELECT COUNT(*) FROM learning_topics WHERE item_id = ?", (item_id,)
            ).fetchone()
            count = existing[0] if not isinstance(existing, dict) else existing['count']
            if count and count > 0:
                return
            for i, name in enumerate(topic_names):
                conn.execute(
                    "INSERT INTO learning_topics (id, item_id, topic_name, order_index, covered) VALUES (?, ?, ?, ?, 0)",
                    (str(uuid.uuid4()), item_id, name, i),
                )
            conn.commit()

    def get_learning_topics(self, item_id: str) -> list:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute(
                "SELECT * FROM learning_topics WHERE item_id = ? ORDER BY order_index ASC", (item_id,)
            ).fetchall()
        return rows

    def toggle_learning_topic(self, topic_id: str) -> bool:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute("SELECT covered FROM learning_topics WHERE id = ?", (topic_id,)).fetchone()
            if not row:
                return False
            new_val = 0 if row['covered'] else 1
            conn.execute("UPDATE learning_topics SET covered = ? WHERE id = ?", (new_val, topic_id))
            conn.commit()
        return bool(new_val)

    # ── Book/PDF library (upload -> extracted text per page, stored in the DB
    # so it survives a restart; the raw PDF itself lives on Cloudinary via
    # cloudinary_url, for the same restart-safety plus a download link) ──────

    def add_book(self, user_id: str, title: str, filename: str, page_texts: list, cloudinary_url: str = "", book_id: str = "") -> str:
        book_id = book_id or str(uuid.uuid4())
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute(
                "INSERT INTO learning_books (id, user_id, title, filename, page_count, current_page, uploaded_at, cloudinary_url) VALUES (?, ?, ?, ?, ?, 1, ?, ?)",
                (book_id, user_id, title, filename, len(page_texts), now, cloudinary_url),
            )
            for i, text in enumerate(page_texts, start=1):
                conn.execute(
                    "INSERT INTO learning_book_pages (id, book_id, page_num, text, summary) VALUES (?, ?, ?, ?, NULL)",
                    (str(uuid.uuid4()), book_id, i, text),
                )
            conn.commit()
        return book_id

    def get_books(self, user_id: str) -> list:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute(
                "SELECT * FROM learning_books WHERE user_id = ? ORDER BY uploaded_at DESC", (user_id,)
            ).fetchall()
        return rows

    def get_book(self, user_id: str, book_id: str) -> dict | None:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute(
                "SELECT * FROM learning_books WHERE id = ? AND user_id = ?", (book_id, user_id)
            ).fetchone()
        return row

    def get_book_page(self, user_id: str, book_id: str, page_num: int) -> dict | None:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute("""
                SELECT p.* FROM learning_book_pages p
                JOIN learning_books b ON b.id = p.book_id
                WHERE p.book_id = ? AND p.page_num = ? AND b.user_id = ?
            """, (book_id, page_num, user_id)).fetchone()
        return row

    def save_page_summary(self, book_id: str, page_num: int, summary: str) -> None:
        with self._get_conn() as conn:
            conn.execute(
                "UPDATE learning_book_pages SET summary = ? WHERE book_id = ? AND page_num = ?",
                (summary, book_id, page_num),
            )
            conn.commit()

    def update_book_current_page(self, user_id: str, book_id: str, page_num: int) -> None:
        with self._get_conn() as conn:
            conn.execute(
                "UPDATE learning_books SET current_page = ? WHERE id = ? AND user_id = ?", (page_num, book_id, user_id)
            )
            conn.commit()

    # ── YouTube playlist study RAG ──────────────────────────────────────────

    def add_playlist(self, user_id: str, url: str, title: str) -> str:
        playlist_id = str(uuid.uuid4())
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute(
                "INSERT INTO learning_playlists (id, user_id, url, title, status, created_at) VALUES (?, ?, ?, ?, 'ingesting', ?)",
                (playlist_id, user_id, url, title, now),
            )
            conn.commit()
        return playlist_id

    def update_playlist_status(self, playlist_id: str, status: str) -> None:
        with self._get_conn() as conn:
            conn.execute("UPDATE learning_playlists SET status = ? WHERE id = ?", (status, playlist_id))
            conn.commit()

    def get_playlists(self, user_id: str) -> list:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute(
                "SELECT * FROM learning_playlists WHERE user_id = ? ORDER BY created_at DESC", (user_id,)
            ).fetchall()
        return rows

    def get_playlist(self, user_id: str, playlist_id: str) -> dict | None:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            row = conn.execute(
                "SELECT * FROM learning_playlists WHERE id = ? AND user_id = ?", (playlist_id, user_id)
            ).fetchone()
        return row

    def add_video(self, playlist_id: str, video_id: str, title: str, url: str) -> str:
        row_id = str(uuid.uuid4())
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute(
                "INSERT INTO learning_videos (id, playlist_id, video_id, title, url, status, created_at) "
                "VALUES (?, ?, ?, ?, ?, 'pending', ?)",
                (row_id, playlist_id, video_id, title, url, now),
            )
            conn.commit()
        return row_id

    def update_video(self, row_id: str, **fields) -> None:
        """fields: any of transcript, notes_md, source, status, error."""
        if not fields:
            return
        cols = ", ".join(f"{k} = ?" for k in fields)
        with self._get_conn() as conn:
            conn.execute(f"UPDATE learning_videos SET {cols} WHERE id = ?", (*fields.values(), row_id))
            conn.commit()

    def get_videos_for_playlist(self, playlist_id: str) -> list:
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute(
                "SELECT * FROM learning_videos WHERE playlist_id = ? ORDER BY created_at", (playlist_id,)
            ).fetchall()
        return rows

    def add_chunks(self, video_row_id: str, playlist_id: str, chunks: list) -> list:
        """chunks: list of chunk text strings. Returns the generated chunk ids,
        in the same order — callers need this order to line up with FAISS rows."""
        ids = [str(uuid.uuid4()) for _ in chunks]
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            for i, (chunk_id, text) in enumerate(zip(ids, chunks)):
                conn.execute(
                    "INSERT INTO learning_video_chunks (id, video_id, playlist_id, chunk_index, text, created_at) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (chunk_id, video_row_id, playlist_id, i, text, now),
                )
            conn.commit()
        return ids

    def get_chunks_by_ids(self, chunk_ids: list) -> dict:
        """Returns {chunk_id: {"text":..., "video_id":..., "playlist_id":..., "video_title":...,
        "owner_user_id":...}}. owner_user_id comes from the chunk's playlist (chunks/videos
        themselves aren't user_id-scoped directly — see tracker.py's schema comment) and lets
        callers filter a shared FAISS index down to one user's own content, e.g. StudyAgent.ask()."""
        if not chunk_ids:
            return {}
        placeholders = ",".join("?" for _ in chunk_ids)
        with self._get_conn() as conn:
            conn.row_factory = ROW_DICT
            rows = conn.execute(
                f"""SELECT c.id, c.text, c.video_id, c.playlist_id, v.title AS video_title,
                           p.user_id AS owner_user_id
                    FROM learning_video_chunks c
                    JOIN learning_videos v ON v.id = c.video_id
                    JOIN learning_playlists p ON p.id = c.playlist_id
                    WHERE c.id IN ({placeholders})""",
                chunk_ids,
            ).fetchall()
        return {r["id"]: r for r in rows}

    # ── Telegram inbound (message_id -> job_id, so a reply to a job alert can
    # be matched back to the job it's about) ──────────────────────────────────

    def record_telegram_alert(self, user_id: str, message_id: str, job_id: str):
        now = datetime.now().isoformat()
        with self._get_conn() as conn:
            conn.execute("""
                INSERT INTO telegram_alerts (message_id, user_id, job_id, sent_at) VALUES (?, ?, ?, ?)
                ON CONFLICT (message_id) DO UPDATE SET job_id = EXCLUDED.job_id
            """, (str(message_id), user_id, job_id, now))
            conn.commit()

    def get_job_id_by_telegram_message(self, message_id: str) -> str | None:
        with self._get_conn() as conn:
            row = conn.execute(
                "SELECT job_id FROM telegram_alerts WHERE message_id = ?", (str(message_id),)
            ).fetchone()
        if not row:
            return None
        return row[0] if not isinstance(row, dict) else row["job_id"]
